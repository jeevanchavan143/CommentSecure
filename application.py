#import modules  Test Video ID: cwzp_F8kOHc 5sLYAQS9sWQ
from tkinter import messagebox
import xlsxwriter
from tkinter import *
from tkinter import ttk
import pandas as pd
from apiclient.discovery import build
from rich import print
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import MultinomialNB
from sklearn.ensemble import VotingClassifier
from sklearn.metrics import accuracy_score


# arguments to be passed to build function
DEVELOPER_KEY = "AIzaSyCReNA7gPbFdimdisbbpH-5Rs9PNluEYig"
YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"


workbook = xlsxwriter.Workbook('demo.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': True})
worksheet.write('A1', 'USERNAME')
worksheet.write('B1', 'PASSWORD')

window = Tk()
window.title("Welcome to CommentSecure :YouTube Spam Detection system")
window.geometry('1200x900')

tab_control = ttk.Notebook(window)
tab1 = ttk.Frame(tab_control)
tab2 = ttk.Frame(tab_control)
tab_control.add(tab1, text='New User Registration')
tab_control.add(tab2, text='Comment Secure')


#############################################################################################################################################################
# HEADING
def show_entry_fields():
    Un = e1.get()
    Pw = e2.get()
    res = "New User " + Un + " is added"
    lbl1.configure(text=res)
    worksheet.write(str('A' + str(2)), str(Un))
    worksheet.write(str('B' + str(2)), str(Pw))
    workbook.close()


def TST_Face():
    Un = ee1.get()
    Pw = ee2.get()
    VI = ee3.get()
    video_id = VI
    print('User Logged in Successfully :', Un)
    import openpyxl
    wb = openpyxl.load_workbook('demo.xlsx')
    sheet = wb.active  # Assuming you want to work with the active sheet

    Un1 = sheet.cell(row=2, column=1).value
    Pw1 = sheet.cell(row=2, column=2).value

    if Un == Un1 and Pw == Pw1:
        lbl211.configure(text="Login Successful")
        lbl20.configure(text="Result")
        lbl21.configure(text="Comment")
    else:
        messagebox.showerror('LOGIN DENIED', 'Wrong Username Or Password')
        window.quit()
        window.destroy()
    # IF LOGIN SUCCESFUL


    # Training data from open source dataaset
    # https://github.com/rahulg-101/Youtube-Comment-Spam-Detection/blob/main/Youtube01.csv
    csv = pd.read_csv('Youtube01.csv')
    comments = csv['CONTENT'].to_list()
    labels = csv['CLASS'].to_list()

    # Splitting the data into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(comments, labels, test_size=0.2, random_state=42)

    # Vectorizing the comments using TF-IDF
    tfidf_vectorizer = TfidfVectorizer(stop_words='english')
    X_train_tfidf = tfidf_vectorizer.fit_transform(X_train)
    X_test_tfidf = tfidf_vectorizer.transform(X_test)

    # Initializing classifiers
    svm_classifier = SVC(kernel='linear', probability=True)
    knn_classifier = KNeighborsClassifier(n_neighbors=5)
    nb_classifier = MultinomialNB()

    # Ensemble classifier using voting
    ensemble_classifier = VotingClassifier(estimators=[
        ('svm', svm_classifier),
        ('knn', knn_classifier),
        ('nb', nb_classifier)
    ], voting='soft')

    # Training individual classifiers
    svm_classifier.fit(X_train_tfidf, y_train)
    knn_classifier.fit(X_train_tfidf, y_train)
    nb_classifier.fit(X_train_tfidf, y_train)
    ensemble_classifier.fit(X_train_tfidf, y_train)

    # Predictions
    svm_pred = svm_classifier.predict(X_test_tfidf)
    knn_pred = knn_classifier.predict(X_test_tfidf)
    nb_pred = nb_classifier.predict(X_test_tfidf)
    ensemble_pred = ensemble_classifier.predict(X_test_tfidf)

    # Accuracy
    svm_accuracy = accuracy_score(y_test, svm_pred)
    knn_accuracy = accuracy_score(y_test, knn_pred)
    nb_accuracy = accuracy_score(y_test, nb_pred)
    ensemble_accuracy = accuracy_score(y_test, ensemble_pred)

    print("SVM Accuracy:", svm_accuracy)
    print("KNN Accuracy:", knn_accuracy)
    print("Naive Bayes Accuracy:", nb_accuracy)
    print("Ensemble Accuracy:", ensemble_accuracy)

    youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION, developerKey=DEVELOPER_KEY)
    #video_id = video_id     # 'cwzp_F8kOHc'  # Change here for different YouTube videos
    max_results = 10
    results = youtube.commentThreads().list(videoId=video_id, part="id,snippet", order="relevance",
                                            textFormat="plainText", maxResults=max_results % 101).execute()
    comments = [
        item['snippet']['topLevelComment']['snippet']['textOriginal']
        for item in results['items']
    ]
    comments_tfidf = tfidf_vectorizer.transform(comments)
    svm_prediction = svm_classifier.predict(comments_tfidf)
    knn_prediction = knn_classifier.predict(comments_tfidf)
    nb_prediction = nb_classifier.predict(comments_tfidf)
    ensemble_prediction = ensemble_classifier.predict(comments_tfidf)

    # Create labels dynamically
    labels = {}
    for i, comment in enumerate(comments):
        label = Label(tab2, text="" , font=("Arial Bold", 10), foreground="red", background="white")
        label.grid(column=0, row=i + 9)  # Adjust column and row as per your requirement
        labels[i] = label

    # Create labels dynamically
    labels1 = {}
    for i, comment in enumerate(comments):
        label = Label(tab2, text="" , font=("Arial Bold", 10), foreground="red", background="white")
        label.grid(column=1, row=i + 9)  # Adjust column and row as per your requirement
        labels1[i] = label

    print('predictions: Comment | SVM | KNN | NB | Ensemble')
    for i, preds in enumerate(zip(comments, svm_prediction, knn_prediction, nb_prediction, ensemble_prediction)):
        print(f'{preds[0]} | {preds[1]} | {preds[2]} | {preds[3]} | {preds[4]}')
        if i in labels:
            labels1[i].configure(text=preds[0])
            if preds[4] == 0:
                labels[i].configure(text="Not SPAM")
            else:
                labels[i].configure(text="SPAM")
        else:
            print(f"Label with index {i} not found in dictionary")
        print("#######################################")

    '''

    print('predictions: Comment | SVM | KNN | NB | Ensemble')
    for i, preds in enumerate(zip(comments, svm_prediction, knn_prediction, nb_prediction, ensemble_prediction)):
        print(f'{preds[0]} | {preds[1]} | {preds[2]} | {preds[3]} | {preds[4]}')
        
        if preds[4] == 0:
            labels[i+10].configure(text="Not SPAM")
        else:
            labels[i+10].configure(text="SPAM")
        labels1[i+20].configure(text=preds[0])
        
        print("#######################################")
        
    '''

'''
    print('predictions: Comment | SVM | KNN | NB | Ensemble')
    for preds in zip(comments, svm_prediction, knn_prediction, nb_prediction, ensemble_prediction):
        print(f'{preds[0]} | {preds[1]} | {preds[2]} | {preds[3]} | {preds[4]}')
        if preds[4]==0:
            lbl23.configure(text="Not SPAM")
        else:
            lbl23.configure(text="SPAM")
        lbl22.configure(text=preds[0])
        print("#######################################")
'''



'''
    # Extracting required info from each result
    for result in results['items']:
        comment = {}
        comments.append(comment)
        u = result['snippet']['topLevelComment']['snippet']['textOriginal']
        u = u.encode('unicode-escape').decode('utf-8')
        print(u)
        RESULT = CLF_fused.classify(u)
        print(RESULT);
        if RESULT == 0:
            RESULT = 'NA'
        elif RESULT == 1:
            RESULT = 'SPAM'
        lbl20.configure(text=u)
        lbl21.configure(text=RESULT)
        time.sleep(2)

'''
#######################################################################################################
lbl = Label(tab1, text="CommentSecure : ", font=("Arial Bold", 30), foreground=("blue"), background=("white"))
lbl.grid(column=0, row=0)
lbl = Label(tab1, text=" Spam", font=("Arial Bold", 30), foreground=("blue"), background=("white"))
lbl.grid(column=1, row=0)
lbl = Label(tab1, text="Classifier", font=("Arial Bold", 30), foreground=("blue"), background=("white"))
lbl.grid(column=2, row=0)
# USERNAME & PASSWORD ENTRY BOX
Label(tab1, text="Username", font=("Arial Bold", 15), foreground=("green")).grid(row=1, column=0)
Label(tab1, text="Password", font=("Arial Bold", 15), foreground=("green")).grid(row=2, column=0)

e1 = Entry(tab1)
e2 = Entry(tab1)

e1.grid(row=1, column=1)
e2.grid(row=2, column=1)


lbl1 = Label(tab1, text="", font=("Arial Bold", 10), foreground=("red"), background=("white"))
lbl1.grid(column=1, row=7)
Button(tab1, text='CANCEL', command=tab1.quit).grid(row=6, column=1, sticky=W, pady=4)
Button(tab1, text='REGISTER', command=show_entry_fields).grid(row=6, column=2, sticky=W, pady=4)
#############################################################################################################################################################
lbl = Label(tab2, text="SPAM", font=("Arial Bold", 30), foreground=("red"), background=("white"))
lbl.grid(column=0, row=0)
lbl = Label(tab2, text="CLASSIFICATION", font=("Arial Bold", 30), foreground=("red"), background=("white"))
lbl.grid(column=1, row=0)
lbl = Label(tab2, text="SYSTEM", font=("Arial Bold", 30), foreground=("red"), background=("white"))
lbl.grid(column=2, row=0)
# USERNAME & PASSWORD ENTRY BOX
Label(tab2, text="USERNAME", font=("Arial Bold", 15), foreground=("green")).grid(row=1, column=0)
Label(tab2, text="PASSWORD", font=("Arial Bold", 15), foreground=("green")).grid(row=2, column=0)
Label(tab2, text="VIDEO ID", font=("Arial Bold", 15), foreground=("green")).grid(row=3, column=0)
ee1 = Entry(tab2)
ee2 = Entry(tab2)
ee3 = Entry(tab2)
ee1.grid(row=1, column=1)
ee2.grid(row=2, column=1)
ee3.grid(row=3, column=1)

lbl211 = Label(tab2, text="", font=("Arial Bold", 10), foreground=("red"), background=("white"))
lbl211.grid(column=1, row=7)
lbl21 = Label(tab2, text="", font=("Arial Bold", 10), foreground=("red"), background=("white"))
lbl21.grid(column=1, row=8)
lbl20 = Label(tab2, text="", font=("Arial Bold", 10), foreground=("red"), background=("white"))
lbl20.grid(column=0, row=8)

#RESULT
'''
labels = []
for i in range(10):
    label = Label(tab2, text="  ", font=("Arial Bold", 10), foreground=("red"), background=("white"))
    label.grid(column=0, row=i+9)  # Adjust column and row as per your requirement
    labels.append(label)

labels1 = []
for i in range(10):
    label = Label(tab2, text="  ", font=("Arial Bold", 10), foreground=("red"), background=("white"))
    label.grid(column=1, row=i+9)  # Adjust column and row as per your requirement
    labels1.append(label)

print(labels)
print(labels1)
'''

from tkinter import Tk, Label

'''
# Assuming you have already initialized Tkinter
# Example values for predictions
comments = ['comment1', 'comment2', 'comment3']
svm_prediction = [0, 1, 0]
knn_prediction = [1, 0, 1]
nb_prediction = [0, 0, 1]
ensemble_prediction = [1, 1, 0]
'''


#lbl22 = Label(tab2, text="", font=("Arial Bold", 10), foreground=("red"), background=("white"))
#lbl22.grid(column=1, row=9)
#lbl23 = Label(tab2, text="", font=("Arial Bold", 10), foreground=("red"), background=("white"))
#lbl23.grid(column=0, row=9)


Button(tab2, text='CANCEL', command=tab2.quit).grid(row=6, column=1, sticky=W, pady=4)
Button(tab2, text='LOGIN', command=TST_Face).grid(row=6, column=2, sticky=W, pady=4)
#############################################################################################################################################################
tab_control.pack(expand=1, fill='both')
window.mainloop()
